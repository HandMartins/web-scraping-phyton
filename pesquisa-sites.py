from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait      #Permite esperar um tempo antes de jogar um erro
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions   # Lança diferetes tipos de erros conforme indicado
import openpyxl
from time import sleep
# import time
# import random
# from datetime import datetime


class PesquisaSites:

    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--lang=pt-BR')
        self.driver = webdriver.Chrome(executable_path=r'C:\Windows\chromedriver.exe',options=chrome_options)
        #self.driver.implicitly_wait(10)
        self.wait = WebDriverWait(
            driver=self.driver,
            timeout=10,
            poll_frequency=1,
            ignored_exceptions=[NoSuchElementException, ElementNotVisibleException, ElementNotSelectableException]
        )
        
        
    def iniciar(self):    
        self.driver.get("https://www.lojasantoantonio.com.br/")
        # self.wait.until(expected_conditions.element_to_be_clickable((By.XPATH, f'//tag[@id="xxx"]')))
        self.pesquisar()
        self.capturaElementos()
        self.criarPlanilha()
        self.armazenarValoresPlanilha()
        print("Finalizando Programa")

     
    def pesquisar(self):
        campo_pesquisar = self.wait.until(expected_conditions.element_to_be_clickable((By.CSS_SELECTOR, f'div[class="linha linha__principal"]')))
        campo_pesquisar = campo_pesquisar.find_element_by_xpath("//input[@class='fulltext-search-box ui-autocomplete-input']")
        campo_pesquisar.send_keys('forma bwb')
        campo_pesquisar.send_keys(Keys.ENTER)
        sleep(3)
        self.expandindoPagina()


    def expandindoPagina(self):
        try:
            while True:
                # sleep(2)
                botao = self.wait.until(expected_conditions.element_to_be_clickable((By.XPATH, f'//button[text()="Carregar Mais"]')))
                botao.click()
                sleep(2)   
        except Exception as erro:
            print("Não ha mais pesquisas")
            print(erro)
            pass


    def capturaElementos(self): 
        self.lista_titulos = self.driver.find_elements_by_xpath('//div[@class="nome"]')
        print(len(self.lista_titulos))
        # for item in self.lista_titulos:
        #     print(item.text)

        self.lista_precos = self.driver.find_elements_by_xpath('//div[@class="principal" or contains(text(), "Produto indisponivel")]')
        print(len(self.lista_precos))
        # for preco in self.lista_precos:
        #     print(preco.text)
           
                
    def criarPlanilha(self):
        self.planilha = openpyxl.Workbook()                             
        self.planilha.create_sheet('Guia_StAntonio')                    
        self.planilha_valores = self.planilha['Guia_StAntonio']         
        self.planilha_valores.cell(row=1,column=1,value='Titulo')
        self.planilha_valores.cell(row=1,column=2,value='Valor')


    def armazenarValoresPlanilha(self):
        try:
            print(len(self.lista_titulos))
            print(len(self.lista_titulos) - 1)
            print(type(len(self.lista_titulos)))
            for indice in range(0, len(self.lista_titulos) - 1):
                nova_linha = [self.lista_titulos[indice].text, self.lista_precos[indice].text]
                self.planilha_valores.append(nova_linha)
            self.planilha.save("DadosLojaSTAntonio.xlsx")
        except Exception as erro:
            print("Erro na classe de armazenarValoresPlanilha")
            print(erro)
            pass
          


curso = PesquisaSites()
curso.iniciar()





















# Aaaah, agora que me toquei, o “R$” vem junto. Nesse caso voce pode fazer isso:
# Vamos dizer que o precos[0].text vale “R$ 4.000”
# preco = precos[0].text # Recebe o texto do elemento escolhido
# preco_sem_cifrao = preco[3:] # Pega os dados a partir do index 3 até o final, tirando “R$” ficando “4.000”
# preco_sem_ponto = preco_sem_cifrao.replace(".","") # Elimina o “.” do numeral, ficando “4000”
# ultimo_preco =int(preco_sem_ponto) #transforma string em numeral